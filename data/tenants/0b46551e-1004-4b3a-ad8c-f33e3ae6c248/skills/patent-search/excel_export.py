#!/usr/bin/env python3
"""检索结果导出 Excel（openpyxl 不可用时回退 CSV）。"""

from __future__ import annotations

import csv
import importlib.util
import io
import os
import re
import uuid
from pathlib import Path
from typing import Any, Sequence

_SKILL_DIR = Path(__file__).resolve().parent
_deps_mod: Any = None

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_MIME = "text/csv; charset=utf-8"


def _load_deps() -> Any:
    global _deps_mod
    if _deps_mod is not None:
        return _deps_mod
    path = _SKILL_DIR / "export_deps.py"
    spec = importlib.util.spec_from_file_location(
        "skill_export_deps_patent_search", path
    )
    if spec is None or spec.loader is None:
        raise ImportError(f"无法加载 {path}")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    _deps_mod = mod
    return mod


def strip_markup(text: Any) -> str:
    if text is None:
        return ""
    s = str(text)
    s = re.sub(r"</?em>", "", s, flags=re.IGNORECASE)
    s = re.sub(r"<[^>]+>", "", s)
    return s.replace("\n", " ").strip()


def build_csv_bytes(
    headers: Sequence[str], rows: Sequence[Sequence[Any]]
) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(list(headers))
    for row in rows:
        writer.writerow(
            [strip_markup(c) if isinstance(c, str) else c for c in row]
        )
    return buf.getvalue().encode("utf-8-sig")


def build_xlsx_bytes(
    headers: Sequence[str], rows: Sequence[Sequence[Any]], sheet_name: str = "数据"
) -> bytes:
    deps = _load_deps()
    ok, err = deps.ensure_openpyxl()
    if not ok:
        raise deps.ExportNotAvailable(err or "openpyxl 不可用")

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "数据"
    ws.append(list(headers))
    for row in rows:
        ws.append([strip_markup(c) if isinstance(c, str) else c for c in row])
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = 12
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 48))
        ws.column_dimensions[letter].width = max_len + 2
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_root() -> Path:
    raw = (os.environ.get("MCHAT_UPLOAD_DIR") or os.environ.get("UPLOAD_DIR") or "").strip()
    if raw:
        return Path(raw).expanduser()
    return Path(__file__).resolve().parents[2] / "uploads"


def save_export_file(
    data: bytes,
    filename: str,
    *,
    subdir: str = "patent-exports",
    mime: str,
    ext: str,
) -> dict[str, str]:
    key = f"{subdir}/{uuid.uuid4().hex}{ext}"
    full = _upload_root() / key
    full.parent.mkdir(parents=True, exist_ok=True)
    full.write_bytes(data)
    display = filename if filename.endswith(ext) else f"{filename}{ext}"
    return {
        "type": "file",
        "name": display,
        "url": f"/uploads/{key}",
        "mime": mime,
    }


def export_table(
    filename: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    sheet_name: str = "数据",
    subdir: str = "patent-exports",
) -> tuple[dict[str, str], str]:
    """
    优先 xlsx，失败则 CSV。返回 (asset, format_note)。
    """
    deps = _load_deps()
    base = filename.replace(".xlsx", "").replace(".csv", "")
    try:
        data = build_xlsx_bytes(headers, rows, sheet_name=sheet_name)
        asset = save_export_file(
            data, base, subdir=subdir, mime=XLSX_MIME, ext=".xlsx"
        )
        return asset, "Excel (.xlsx)"
    except deps.ExportNotAvailable:
        data = build_csv_bytes(headers, rows)
        asset = save_export_file(
            data, base, subdir=subdir, mime=CSV_MIME, ext=".csv"
        )
        return (
            asset,
            "CSV（Excel 依赖未就绪，已自动改用 CSV，仍可用 Excel/WPS 打开）",
        )


# 兼容旧调用
def save_xlsx_file(
    data: bytes,
    filename: str,
    *,
    subdir: str = "exports",
) -> dict[str, str]:
    return save_export_file(
        data, filename, subdir=subdir, mime=XLSX_MIME, ext=".xlsx"
    )

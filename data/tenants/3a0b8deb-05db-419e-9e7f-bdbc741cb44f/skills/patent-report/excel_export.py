"""Export sections to Excel."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from io_utils import file_artifact, sanitize_filename
from openpyxl import Workbook
from openpyxl.styles import Font


def _sheet_title(title: str, used: set[str]) -> str:
    base = sanitize_filename(title, default="Sheet")[:31] or "Sheet"
    name = base
    idx = 1
    while name in used:
        suffix = f"_{idx}"
        name = (base[: 31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        idx += 1
    used.add(name)
    return name


def export_excel(
    sections: list[dict[str, Any]],
    *,
    out_dir: Path,
    key_prefix: str,
    title: str,
    filename: str,
    summary: str = "",
    interpretation: str = "",
) -> dict[str, Any]:
    wb = Workbook()
    overview = wb.active
    overview.title = "Summary"
    overview["A1"] = title
    overview["A1"].font = Font(bold=True, size=14)
    overview["A2"] = "Section"
    overview["B2"] = "Rows"
    overview["A2"].font = Font(bold=True)
    overview["B2"].font = Font(bold=True)

    used_titles = {"Summary"}
    if summary or interpretation:
        narrative = wb.create_sheet("总结与解读")
        used_titles.add("总结与解读")
        narrative["A1"] = "总结"
        narrative["A1"].font = Font(bold=True, size=12)
        narrative["A2"] = summary
        narrative["A4"] = "解读"
        narrative["A4"].font = Font(bold=True, size=12)
        narrative["A5"] = interpretation

    row_idx = 3
    for section in sections:
        overview[f"A{row_idx}"] = section.get("title") or "Section"
        overview[f"B{row_idx}"] = len(section.get("rows") or [])
        row_idx += 1

        ws = wb.create_sheet(_sheet_title(str(section.get("title") or "Section"), used_titles))
        ws.append(["Label", "Value"])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for item in section.get("rows") or []:
            ws.append([item.get("label"), item.get("value")])
        if not section.get("rows") and section.get("text"):
            ws.append(["Notes"])
            ws.append([section.get("text")])

    fname = sanitize_filename(filename, default="patent-report") + ".xlsx"
    path = out_dir / fname
    wb.save(path)
    key = f"{key_prefix}/{fname}"
    return file_artifact(path, key, fmt="xlsx")
